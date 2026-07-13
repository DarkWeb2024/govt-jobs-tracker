"""Multi-sheet Excel workbook output."""
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .tabular import FIELDS

HDR_FILL = PatternFill("solid", start_color="1F4E78")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY = Font(name="Arial", size=10)
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
STATUS_FILLS = {
    "Critical": PatternFill("solid", start_color="FFC7CE"),
    "High": PatternFill("solid", start_color="FFE2C7"),
    "Medium": PatternFill("solid", start_color="FFF2CC"),
}


def _sheet(wb, title, items):
    ws = wb.create_sheet(title[:31])
    headers = ["days_left"] + FIELDS
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, THIN
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for n in items:
        d = n.to_dict()
        ws.append([n.days_left()] + [d[f] for f in FIELDS])
        r = ws.max_row
        fill = STATUS_FILLS.get(n.priority)
        for i in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=i)
            cell.font, cell.border = BODY, THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
        for col, field in ((19, "official_website"), (20, "official_pdf"), (21, "apply_link")):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, str) and v.startswith("http"):
                ws.cell(row=r, column=col).hyperlink = v
                ws.cell(row=r, column=col).font = Font(name="Arial", size=10,
                                                       color="0563C1", underline="single")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 2)}"
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 42


def _history_sheet(wb, history):
    ws = wb.create_sheet("History")
    ws.append(["changed_at", "notification_id", "field", "old_value", "new_value"])
    for c in ws[1]:
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, THIN
    for nid, ts, field, old, new in history:
        ws.append([ts, nid, field, str(old)[:120], str(new)[:120]])
    for col, w in zip("ABCDE", (20, 16, 22, 40, 40)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(ws.max_row, 2)}"


def write_excel(notifications, path, history=None):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    verified = [n for n in notifications if n.verification_status.startswith("Verified")]
    unverified = [n for n in notifications if n.verification_status == "Unverified"]
    from ..models import is_applied
    applied = [n for n in notifications if is_applied(n.status)]
    open_now = [n for n in notifications
                if (n.days_left() is None or n.days_left() >= 0)
                and n.verification_status not in ("Application Closed", "Expired", "Cancelled")]
    upcoming = sorted([n for n in open_now if n.days_left() is not None],
                      key=lambda n: n.days_left())
    expired = [n for n in notifications
               if n.verification_status in ("Application Closed", "Expired")]
    jobs = [n for n in notifications if n.category != "Exam"]
    exams = [n for n in notifications if n.category == "Exam"]

    def _is_karnataka(n):
        return "karnataka" in f"{n.state} {n.location}".lower()
    karnataka = [n for n in notifications if _is_karnataka(n)]
    central = [n for n in notifications if not _is_karnataka(n)]
    from ..models import HIDDEN_STATES
    not_interested = [n for n in notifications if n.status in HIDDEN_STATES]

    for title, items in [("Verified", verified), ("Unverified", unverified),
                         ("Applied", applied), ("Jobs", jobs), ("Exams", exams),
                         ("Central", central), ("Karnataka", karnataka),
                         ("Not Interested", not_interested),
                         ("Expired", expired), ("Upcoming", upcoming),
                         ("Calendar", upcoming)]:
        _sheet(wb, title, items)
    if history:
        _history_sheet(wb, history)

    stats = wb.create_sheet("Statistics", 0)
    stats["A1"] = "Statistics"
    stats["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E78")
    rows = [
        ("Total tracked", len(notifications)),
        ("Verified", len(verified)), ("Unverified", len(unverified)),
        ("Applied", len(applied)), ("Open now", len(open_now)),
        ("Expired/Closed", len(expired)),
        ("Critical priority", sum(1 for n in notifications if n.priority == "Critical")),
        ("High priority", sum(1 for n in notifications if n.priority == "High")),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        stats[f"A{i}"], stats[f"B{i}"] = k, v
        stats[f"A{i}"].font = BODY
        stats[f"B{i}"].font = Font(name="Arial", size=10, bold=True)
    stats.column_dimensions["A"].width = 28
    wb.save(path)
    return path
